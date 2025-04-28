
#!/usr/bin/env python3
"""
Excel Data Import Tool

This script imports data from Excel files for local development setup.

Requirements:
- Python 3.6+
- openpyxl

Usage:
python import_excel_data.py <excel_file_path>
"""

import sys
import os
import openpyxl
from datetime import datetime
import json

def import_departments(wb, output_path):
    """Import departments from Excel worksheet."""
    try:
        ws = wb["Departments"]
        output_wb = openpyxl.Workbook()
        output_ws = output_wb.active
        
        # Copy headers
        for col in range(1, ws.max_column + 1):
            output_ws.cell(row=1, column=col, value=ws.cell(row=1, column=col).value)
        
        # Copy data
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                output_ws.cell(row=row, column=col, value=ws.cell(row=row, column=col).value)
        
        output_file = os.path.join(output_path, 'departments.xlsx')
        output_wb.save(output_file)
        print(f"Successfully imported departments to {output_file}")
        
        # Create a default department if none exists
        if ws.max_row <= 1:
            output_ws.cell(row=2, column=1, value="IT")
            output_ws.cell(row=2, column=2, value="Information Technology")
            output_ws.cell(row=3, column=1, value="HR")
            output_ws.cell(row=3, column=2, value="Human Resources")
            output_wb.save(output_file)
            print("Created default departments")
    except Exception as e:
        print(f"Error importing departments: {e}")

def import_users(wb, output_path):
    """Import users from Excel worksheet."""
    try:
        ws = wb["Users"]
        output_wb = openpyxl.Workbook()
        output_ws = output_wb.active
        
        # Copy headers
        for col in range(1, ws.max_column + 1):
            output_ws.cell(row=1, column=col, value=ws.cell(row=1, column=col).value)
        
        # Copy data
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                output_ws.cell(row=row, column=col, value=ws.cell(row=row, column=col).value)
        
        output_file = os.path.join(output_path, 'users.xlsx')
        output_wb.save(output_file)
        print(f"Successfully imported users to {output_file}")
        
        # Create default users if none exists
        if ws.max_row <= 1:
            # Create headers if they don't exist
            headers = ["username", "password", "fullName", "email", "role", "department", "phone"]
            for col, header in enumerate(headers, start=1):
                output_ws.cell(row=1, column=col, value=header)
                
            # Add admin user
            output_ws.cell(row=2, column=1, value="admin")
            output_ws.cell(row=2, column=2, value="admin123")
            output_ws.cell(row=2, column=3, value="Administrator")
            output_ws.cell(row=2, column=4, value="admin@example.com")
            output_ws.cell(row=2, column=5, value="admin")
            output_ws.cell(row=2, column=6, value="IT")
            output_ws.cell(row=2, column=7, value="123-456-7890")
            
            # Add client user
            output_ws.cell(row=3, column=1, value="client")
            output_ws.cell(row=3, column=2, value="client123")
            output_ws.cell(row=3, column=3, value="Client User")
            output_ws.cell(row=3, column=4, value="client@example.com")
            output_ws.cell(row=3, column=5, value="client")
            output_ws.cell(row=3, column=6, value="HR")
            output_ws.cell(row=3, column=7, value="123-456-7891")
            
            output_wb.save(output_file)
            print("Created default users")
    except Exception as e:
        print(f"Error importing users: {e}")

def import_requests(wb, output_path):
    """Import requests from Excel worksheet."""
    try:
        ws = wb["Requests"]
        output_wb = openpyxl.Workbook()
        output_ws = output_wb.active
        
        # Copy headers
        for col in range(1, ws.max_column + 1):
            output_ws.cell(row=1, column=col, value=ws.cell(row=1, column=col).value)
        
        # Copy data
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                output_ws.cell(row=row, column=col, value=ws.cell(row=row, column=col).value)
        
        output_file = os.path.join(output_path, 'requests.xlsx')
        output_wb.save(output_file)
        print(f"Successfully imported requests to {output_file}")
        
        # Create a sample request if none exists
        if ws.max_row <= 1:
            # Create headers if they don't exist
            headers = ["id", "title", "description", "department", "departments", "status", "creator", "createdAt", 
                      "type", "multiDepartment", "acceptedBy", "usersAccepted", "usersNeeded", "archived", "archivedAt"]
            for col, header in enumerate(headers, start=1):
                output_ws.cell(row=1, column=col, value=header)
                
            # Add sample request
            now = datetime.now().isoformat()
            output_ws.cell(row=2, column=1, value="#100001")
            output_ws.cell(row=2, column=2, value="Sample Request")
            output_ws.cell(row=2, column=3, value="This is a sample request for testing")
            output_ws.cell(row=2, column=4, value="IT")
            output_ws.cell(row=2, column=5, value=json.dumps(["IT"]))
            output_ws.cell(row=2, column=6, value="Pending")
            output_ws.cell(row=2, column=7, value="admin")
            output_ws.cell(row=2, column=8, value=now)
            output_ws.cell(row=2, column=9, value="request")
            output_ws.cell(row=2, column=10, value=False)
            output_ws.cell(row=2, column=11, value=json.dumps([]))
            output_ws.cell(row=2, column=12, value=0)
            output_ws.cell(row=2, column=13, value=1)
            output_ws.cell(row=2, column=14, value=False)
            output_ws.cell(row=2, column=15, value=None)
            
            output_wb.save(output_file)
            print("Created sample request")
    except Exception as e:
        print(f"Error importing requests: {e}")

def main():
    if len(sys.argv) != 2:
        print("Usage: python import_excel_data.py <excel_file_path>")
        sys.exit(1)

    excel_file = sys.argv[1]
    output_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'excel')
    
    print(f"Excel input file: {excel_file}")
    print(f"Output path: {output_path}")
    
    if not os.path.exists(excel_file):
        print(f"Error: File {excel_file} does not exist.")
        # Create a default Excel file
        create_default_excel(excel_file)
    
    if not os.path.exists(output_path):
        os.makedirs(output_path)
        print(f"Created output directory: {output_path}")

    try:
        # Load Excel workbook
        print(f"Loading Excel file: {excel_file}")
        wb = openpyxl.load_workbook(excel_file)
        
        # Check if required sheets exist, if not create them
        required_sheets = ["Departments", "Users", "Requests"]
        for sheet_name in required_sheets:
            if sheet_name not in wb.sheetnames:
                print(f"Sheet {sheet_name} not found, creating...")
                wb.create_sheet(sheet_name)
        
        # Import data
        import_departments(wb, output_path)
        import_users(wb, output_path)
        import_requests(wb, output_path)
        
        print("\nImport completed successfully!")
        print(f"Data files have been created in: {output_path}")

    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)

def create_default_excel(file_path):
    """Create a default Excel file with required sheets"""
    try:
        wb = openpyxl.Workbook()
        
        # Create departments sheet
        dept_ws = wb.active
        dept_ws.title = "Departments"
        dept_ws['A1'] = "id"
        dept_ws['B1'] = "name"
        
        # Create users sheet
        users_ws = wb.create_sheet("Users")
        users_ws['A1'] = "username"
        users_ws['B1'] = "password"
        users_ws['C1'] = "fullName"
        users_ws['D1'] = "email"
        users_ws['E1'] = "role"
        users_ws['F1'] = "department"
        users_ws['G1'] = "phone"
        
        # Create requests sheet
        requests_ws = wb.create_sheet("Requests")
        requests_ws['A1'] = "id"
        requests_ws['B1'] = "title"
        requests_ws['C1'] = "description"
        requests_ws['D1'] = "department"
        requests_ws['E1'] = "departments"
        requests_ws['F1'] = "status"
        requests_ws['G1'] = "creator"
        requests_ws['H1'] = "createdAt"
        requests_ws['I1'] = "type"
        requests_ws['J1'] = "multiDepartment"
        requests_ws['K1'] = "acceptedBy"
        requests_ws['L1'] = "usersAccepted"
        requests_ws['M1'] = "usersNeeded"
        requests_ws['N1'] = "archived"
        requests_ws['O1'] = "archivedAt"
        
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        wb.save(file_path)
        print(f"Created default Excel file at {file_path}")
    except Exception as e:
        print(f"Error creating default Excel file: {e}")

if __name__ == "__main__":
    main()
