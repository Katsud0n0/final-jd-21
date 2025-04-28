
#!/usr/bin/env python3
"""
Excel Database Operations

This script handles all Excel database operations for the application.
It provides a CLI interface for the server to interact with Excel files.

Requirements:
- Python 3.6+
- openpyxl
"""

import sys
import os
import json
import uuid
from datetime import datetime
import openpyxl

# Define the base directory for Excel files
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_DIR = os.path.join(BASE_DIR, 'data', 'excel')

# Ensure the Excel directory exists
if not os.path.exists(EXCEL_DIR):
    os.makedirs(EXCEL_DIR)

# File paths
DEPARTMENTS_FILE = os.path.join(EXCEL_DIR, 'departments.xlsx')
USERS_FILE = os.path.join(EXCEL_DIR, 'users.xlsx')
REQUESTS_FILE = os.path.join(EXCEL_DIR, 'requests.xlsx')

def get_column_names(worksheet):
    """Get column names from the first row of a worksheet."""
    return [cell.value for cell in worksheet[1]]

def rows_to_dicts(worksheet):
    """Convert worksheet rows to list of dictionaries."""
    columns = get_column_names(worksheet)
    result = []
    
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        item = {}
        for i, value in enumerate(row):
            if i < len(columns):
                item[columns[i]] = value
        if any(item.values()):  # Skip empty rows
            result.append(item)
    
    return result

def get_departments():
    """Get all departments from Excel."""
    try:
        if not os.path.exists(DEPARTMENTS_FILE):
            return []
        
        wb = openpyxl.load_workbook(DEPARTMENTS_FILE)
        ws = wb.active
        departments = rows_to_dicts(ws)
        
        return departments
    except Exception as e:
        print(f"Error getting departments: {str(e)}", file=sys.stderr)
        return []

def get_users():
    """Get all users from Excel."""
    try:
        if not os.path.exists(USERS_FILE):
            return []
        
        wb = openpyxl.load_workbook(USERS_FILE)
        ws = wb.active
        users = rows_to_dicts(ws)
        
        # Remove sensitive information
        for user in users:
            if 'password' in user:
                del user['password']
        
        return users
    except Exception as e:
        print(f"Error getting users: {str(e)}", file=sys.stderr)
        return []

def login_user(username, password):
    """Authenticate user by username and password."""
    try:
        if not os.path.exists(USERS_FILE):
            return None
        
        wb = openpyxl.load_workbook(USERS_FILE)
        ws = wb.active
        users = rows_to_dicts(ws)
        
        for user in users:
            if user.get('username') == username:
                # For demo purposes, allow login without password check
                if not password or user.get('password') == password:
                    if 'password' in user:
                        del user['password']
                    return user
        
        return None
    except Exception as e:
        print(f"Error during login: {str(e)}", file=sys.stderr)
        return None

def update_user(user_id, user_data):
    """Update user information in Excel."""
    try:
        if not os.path.exists(USERS_FILE):
            return None
        
        user_data = json.loads(user_data)
        
        wb = openpyxl.load_workbook(USERS_FILE)
        ws = wb.active
        columns = get_column_names(ws)
        
        # Find user row by ID
        user_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == user_id:
                user_row = row_idx
                break
        
        if not user_row:
            return None
        
        # Update user data
        for col_idx, col_name in enumerate(columns, start=1):
            if col_name in user_data:
                ws.cell(row=user_row, column=col_idx, value=user_data[col_name])
        
        wb.save(USERS_FILE)
        
        # Return updated user
        updated_user = {col_name: ws.cell(row=user_row, column=col_idx+1).value 
                      for col_idx, col_name in enumerate(columns)}
        
        if 'password' in updated_user:
            del updated_user['password']
        
        return updated_user
    except Exception as e:
        print(f"Error updating user: {str(e)}", file=sys.stderr)
        return None

def get_requests():
    """Get all requests from Excel."""
    try:
        if not os.path.exists(REQUESTS_FILE):
            return []
        
        wb = openpyxl.load_workbook(REQUESTS_FILE)
        ws = wb.active
        requests = rows_to_dicts(ws)
        
        # Process JSON fields
        for request in requests:
            # Convert string fields to proper types
            for field in ['acceptedBy', 'departments', 'rejections', 'participantsCompleted']:
                if field in request and request[field] and isinstance(request[field], str):
                    try:
                        request[field] = json.loads(request[field])
                    except:
                        request[field] = []
            
            # Convert boolean fields
            for field in ['multiDepartment', 'archived']:
                if field in request and request[field] in ['TRUE', 'True', 'true', 1]:
                    request[field] = True
                elif field in request and request[field] in ['FALSE', 'False', 'false', 0, None, '']:
                    request[field] = False
            
            # Convert numeric fields
            for field in ['usersNeeded', 'usersAccepted']:
                if field in request and request[field] not in [None, '']:
                    try:
                        request[field] = int(request[field])
                    except:
                        pass
        
        return requests
    except Exception as e:
        print(f"Error getting requests: {str(e)}", file=sys.stderr)
        return []

def create_request(request_data):
    """Create a new request in Excel."""
    try:
        request_data = json.loads(request_data)
        
        # Generate ID if not provided
        if 'id' not in request_data:
            request_data['id'] = f"#{uuid.uuid4().hex[:6].upper()}"
        
        # Set created timestamp
        now = datetime.now()
        if 'dateCreated' not in request_data:
            request_data['dateCreated'] = now.strftime("%d/%m/%Y")
        if 'createdAt' not in request_data:
            request_data['createdAt'] = now.isoformat()
        
        # Convert complex fields to strings for Excel storage
        for field in ['acceptedBy', 'departments', 'rejections', 'participantsCompleted']:
            if field in request_data and isinstance(request_data[field], (list, dict)):
                request_data[field] = json.dumps(request_data[field])
        
        # Create request file if it doesn't exist
        if not os.path.exists(REQUESTS_FILE):
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Define columns
            columns = [
                'id', 'title', 'description', 'department', 'status', 'dateCreated', 
                'creator', 'type', 'multiDepartment', 'usersNeeded', 'archived', 
                'archivedAt', 'acceptedBy', 'usersAccepted', 'departments', 
                'rejections', 'participantsCompleted', 'createdAt', 'creatorDepartment',
                'creatorRole', 'lastStatusUpdate', 'lastStatusUpdateTime', 'priority',
                'relatedProject'
            ]
            
            # Write header row
            for col_idx, col_name in enumerate(columns, start=1):
                ws.cell(row=1, column=col_idx, value=col_name)
            
            # Write first data row
            for col_idx, col_name in enumerate(columns, start=1):
                if col_name in request_data:
                    ws.cell(row=2, column=col_idx, value=request_data[col_name])
            
            wb.save(REQUESTS_FILE)
            
        else:
            # Append to existing file
            wb = openpyxl.load_workbook(REQUESTS_FILE)
            ws = wb.active
            columns = get_column_names(ws)
            
            # Find next empty row
            next_row = ws.max_row + 1
            
            # Write data
            for col_idx, col_name in enumerate(columns, start=1):
                if col_name in request_data:
                    ws.cell(row=next_row, column=col_idx, value=request_data[col_name])
            
            wb.save(REQUESTS_FILE)
        
        # Return created request with parsed fields
        for field in ['acceptedBy', 'departments', 'rejections', 'participantsCompleted']:
            if field in request_data and isinstance(request_data[field], str):
                try:
                    request_data[field] = json.loads(request_data[field])
                except:
                    request_data[field] = []
        
        return request_data
    except Exception as e:
        print(f"Error creating request: {str(e)}", file=sys.stderr)
        return None

def update_request(request_id, request_data):
    """Update a request in Excel."""
    try:
        if not os.path.exists(REQUESTS_FILE):
            return None
        
        request_data = json.loads(request_data)
        
        wb = openpyxl.load_workbook(REQUESTS_FILE)
        ws = wb.active
        columns = get_column_names(ws)
        
        # Find request row by ID
        request_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if str(row[0]) == str(request_id):
                request_row = row_idx
                break
        
        if not request_row:
            return None
        
        # Convert complex fields to strings for Excel storage
        for field in ['acceptedBy', 'departments', 'rejections', 'participantsCompleted']:
            if field in request_data and isinstance(request_data[field], (list, dict)):
                request_data[field] = json.dumps(request_data[field])
        
        # Update request data
        for col_idx, col_name in enumerate(columns, start=1):
            if col_name in request_data:
                ws.cell(row=request_row, column=col_idx, value=request_data[col_name])
        
        wb.save(REQUESTS_FILE)
        
        # Return updated request
        updated_request = {col_name: ws.cell(row=request_row, column=col_idx+1).value 
                         for col_idx, col_name in enumerate(columns)}
        
        # Process JSON fields
        for field in ['acceptedBy', 'departments', 'rejections', 'participantsCompleted']:
            if field in updated_request and updated_request[field] and isinstance(updated_request[field], str):
                try:
                    updated_request[field] = json.loads(updated_request[field])
                except:
                    updated_request[field] = []
        
        # Convert boolean fields
        for field in ['multiDepartment', 'archived']:
            if field in updated_request and updated_request[field] in ['TRUE', 'True', 'true', 1]:
                updated_request[field] = True
            elif field in updated_request and updated_request[field] in ['FALSE', 'False', 'false', 0, None, '']:
                updated_request[field] = False
        
        # Convert numeric fields
        for field in ['usersNeeded', 'usersAccepted']:
            if field in updated_request and updated_request[field] not in [None, '']:
                try:
                    updated_request[field] = int(updated_request[field])
                except:
                    pass
        
        return updated_request
    except Exception as e:
        print(f"Error updating request: {str(e)}", file=sys.stderr)
        return None

def delete_request(request_id):
    """Delete a request from Excel."""
    try:
        if not os.path.exists(REQUESTS_FILE):
            return False
        
        wb = openpyxl.load_workbook(REQUESTS_FILE)
        ws = wb.active
        
        # Find request row by ID
        request_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if str(row[0]) == str(request_id):
                request_row = row_idx
                break
        
        if not request_row:
            return False
        
        # Delete row
        ws.delete_rows(request_row, 1)
        
        wb.save(REQUESTS_FILE)
        return True
    except Exception as e:
        print(f"Error deleting request: {str(e)}", file=sys.stderr)
        return False

def accept_request(request_id, username):
    """Accept a request by adding user to acceptedBy."""
    try:
        if not os.path.exists(REQUESTS_FILE):
            return None
        
        wb = openpyxl.load_workbook(REQUESTS_FILE)
        ws = wb.active
        columns = get_column_names(ws)
        column_indices = {name: idx+1 for idx, name in enumerate(columns)}
        
        # Find request row by ID
        request_row = None
        request_data = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if str(row[0].value) == str(request_id):
                request_row = row_idx
                request_data = {col: row[i].value for i, col in enumerate(columns)}
                break
        
        if not request_row or not request_data:
            return None
        
        # Get current acceptedBy list
        accepted_by_str = request_data.get('acceptedBy', '[]')
        try:
            accepted_by = json.loads(accepted_by_str) if accepted_by_str else []
        except:
            accepted_by = []
        
        if not isinstance(accepted_by, list):
            accepted_by = []
        
        # Add user if not already in the list
        if username not in accepted_by:
            accepted_by.append(username)
        
        # Update users accepted count
        users_accepted = int(request_data.get('usersAccepted', 0) or 0) + 1
        users_needed = int(request_data.get('usersNeeded', 1) or 1)
        
        # Check if status should be updated
        status = request_data.get('status', 'Pending')
        if users_accepted >= users_needed and status == 'Pending':
            status = 'In Process'
            
            # Update status timestamp
            now = datetime.now()
            ws.cell(row=request_row, column=column_indices.get('lastStatusUpdate', 0), 
                   value=now.isoformat() if 'lastStatusUpdate' in column_indices else None)
            ws.cell(row=request_row, column=column_indices.get('lastStatusUpdateTime', 0), 
                   value=now.strftime("%H:%M:%S") if 'lastStatusUpdateTime' in column_indices else None)
        
        # Update cells in the worksheet
        ws.cell(row=request_row, column=column_indices.get('acceptedBy', 0), 
               value=json.dumps(accepted_by) if 'acceptedBy' in column_indices else None)
        ws.cell(row=request_row, column=column_indices.get('usersAccepted', 0), 
               value=users_accepted if 'usersAccepted' in column_indices else None)
        ws.cell(row=request_row, column=column_indices.get('status', 0), 
               value=status if 'status' in column_indices else None)
        
        wb.save(REQUESTS_FILE)
        
        # Return updated request
        updated_request = rows_to_dicts(ws)[request_row - 2]
        
        # Process JSON fields
        for field in ['acceptedBy', 'departments', 'rejections', 'participantsCompleted']:
            if field in updated_request and updated_request[field] and isinstance(updated_request[field], str):
                try:
                    updated_request[field] = json.loads(updated_request[field])
                except:
                    updated_request[field] = []
        
        return updated_request
    except Exception as e:
        print(f"Error accepting request: {str(e)}", file=sys.stderr)
        return None

def complete_request(request_id, username):
    """Complete a request."""
    try:
        if not os.path.exists(REQUESTS_FILE):
            return None
        
        wb = openpyxl.load_workbook(REQUESTS_FILE)
        ws = wb.active
        columns = get_column_names(ws)
        column_indices = {name: idx+1 for idx, name in enumerate(columns)}
        
        # Find request row by ID
        request_row = None
        request_data = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if str(row[0].value) == str(request_id):
                request_row = row_idx
                request_data = {col: row[i].value for i, col in enumerate(columns)}
                break
        
        if not request_row or not request_data:
            return None
        
        # Get multi-department status
        multi_department = request_data.get('multiDepartment')
        request_type = request_data.get('type', 'request')
        
        if (multi_department in ['TRUE', 'True', 'true', True, 1] or 
            request_type in ['project', 'Project']):
            # For multi-department requests or projects, track participants who completed
            participants_completed_str = request_data.get('participantsCompleted', '[]')
            try:
                participants_completed = json.loads(participants_completed_str) if participants_completed_str else []
            except:
                participants_completed = []
            
            if not isinstance(participants_completed, list):
                participants_completed = []
            
            # Add user if not already marked as completed
            if username not in participants_completed:
                participants_completed.append(username)
            
            # Get accepted by list
            accepted_by_str = request_data.get('acceptedBy', '[]')
            try:
                accepted_by = json.loads(accepted_by_str) if accepted_by_str else []
            except:
                accepted_by = []
            
            # Check if all participants have completed
            status = request_data.get('status', 'In Process')
            if (len(participants_completed) >= len(accepted_by) and 
                len(accepted_by) >= 2 and status == 'In Process'):
                status = 'Completed'
                
                # Update status timestamp
                now = datetime.now()
                ws.cell(row=request_row, column=column_indices.get('lastStatusUpdate', 0), 
                       value=now.isoformat() if 'lastStatusUpdate' in column_indices else None)
                ws.cell(row=request_row, column=column_indices.get('lastStatusUpdateTime', 0), 
                       value=now.strftime("%H:%M:%S") if 'lastStatusUpdateTime' in column_indices else None)
            
            # Update cells
            ws.cell(row=request_row, column=column_indices.get('participantsCompleted', 0), 
                   value=json.dumps(participants_completed) if 'participantsCompleted' in column_indices else None)
            ws.cell(row=request_row, column=column_indices.get('status', 0), 
                   value=status if 'status' in column_indices else None)
        else:
            # For regular requests, just mark as completed
            now = datetime.now()
            ws.cell(row=request_row, column=column_indices.get('status', 0), 
                   value='Completed' if 'status' in column_indices else None)
            ws.cell(row=request_row, column=column_indices.get('lastStatusUpdate', 0), 
                   value=now.isoformat() if 'lastStatusUpdate' in column_indices else None)
            ws.cell(row=request_row, column=column_indices.get('lastStatusUpdateTime', 0), 
                   value=now.strftime("%H:%M:%S") if 'lastStatusUpdateTime' in column_indices else None)
        
        wb.save(REQUESTS_FILE)
        
        # Return updated request
        updated_request = rows_to_dicts(ws)[request_row - 2]
        
        # Process JSON fields
        for field in ['acceptedBy', 'departments', 'rejections', 'participantsCompleted']:
            if field in updated_request and updated_request[field] and isinstance(updated_request[field], str):
                try:
                    updated_request[field] = json.loads(updated_request[field])
                except:
                    updated_request[field] = []
        
        return updated_request
    except Exception as e:
        print(f"Error completing request: {str(e)}", file=sys.stderr)
        return None

def abandon_request(request_id, username):
    """Abandon a request."""
    try:
        if not os.path.exists(REQUESTS_FILE):
            return None
        
        wb = openpyxl.load_workbook(REQUESTS_FILE)
        ws = wb.active
        columns = get_column_names(ws)
        column_indices = {name: idx+1 for idx, name in enumerate(columns)}
        
        # Find request row by ID
        request_row = None
        request_data = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if str(row[0].value) == str(request_id):
                request_row = row_idx
                request_data = {col: row[i].value for i, col in enumerate(columns)}
                break
        
        if not request_row or not request_data:
            return None
        
        # Get multi-department status
        multi_department = request_data.get('multiDepartment')
        request_type = request_data.get('type', 'request')
        
        now = datetime.now()
        
        if (multi_department in ['TRUE', 'True', 'true', True, 1] or 
            request_type in ['project', 'Project']):
            # For multi-department requests or projects, remove user from participants
            accepted_by_str = request_data.get('acceptedBy', '[]')
            try:
                accepted_by = json.loads(accepted_by_str) if accepted_by_str else []
            except:
                accepted_by = []
            
            if not isinstance(accepted_by, list):
                accepted_by = []
            
            # Remove user if in the list
            if username in accepted_by:
                accepted_by.remove(username)
            
            # Get participants completed list
            participants_completed_str = request_data.get('participantsCompleted', '[]')
            try:
                participants_completed = json.loads(participants_completed_str) if participants_completed_str else []
            except:
                participants_completed = []
            
            # Remove user from completed list
            if username in participants_completed:
                participants_completed.remove(username)
            
            # Update users accepted count
            users_accepted = max(int(request_data.get('usersAccepted', 0) or 0) - 1, 0)
            
            # Add rejection record
            rejections_str = request_data.get('rejections', '[]')
            try:
                rejections = json.loads(rejections_str) if rejections_str else []
            except:
                rejections = []
            
            if not isinstance(rejections, list):
                rejections = []
            
            rejections.append({
                'username': username,
                'reason': '',
                'date': now.strftime("%d/%m/%Y %H:%M:%S")
            })
            
            # Set status back to Pending
            status = 'Pending'
            
            # Update cells
            ws.cell(row=request_row, column=column_indices.get('acceptedBy', 0), 
                   value=json.dumps(accepted_by) if 'acceptedBy' in column_indices else None)
            ws.cell(row=request_row, column=column_indices.get('participantsCompleted', 0), 
                   value=json.dumps(participants_completed) if 'participantsCompleted' in column_indices else None)
            ws.cell(row=request_row, column=column_indices.get('usersAccepted', 0), 
                   value=users_accepted if 'usersAccepted' in column_indices else None)
            ws.cell(row=request_row, column=column_indices.get('status', 0), 
                   value=status if 'status' in column_indices else None)
            ws.cell(row=request_row, column=column_indices.get('rejections', 0), 
                   value=json.dumps(rejections) if 'rejections' in column_indices else None)
            ws.cell(row=request_row, column=column_indices.get('lastStatusUpdate', 0), 
                   value=now.isoformat() if 'lastStatusUpdate' in column_indices else None)
            ws.cell(row=request_row, column=column_indices.get('lastStatusUpdateTime', 0), 
                   value=now.strftime("%H:%M:%S") if 'lastStatusUpdateTime' in column_indices else None)
        else:
            # For regular requests, mark as rejected
            rejections_str = request_data.get('rejections', '[]')
            try:
                rejections = json.loads(rejections_str) if rejections_str else []
            except:
                rejections = []
            
            if not isinstance(rejections, list):
                rejections = []
            
            rejections.append({
                'username': username,
                'reason': '',
                'date': now.strftime("%d/%m/%Y %H:%M:%S")
            })
            
            # Update cells
            ws.cell(row=request_row, column=column_indices.get('status', 0), 
                   value='Rejected' if 'status' in column_indices else None)
            ws.cell(row=request_row, column=column_indices.get('acceptedBy', 0), 
                   value='[]' if 'acceptedBy' in column_indices else None)
            ws.cell(row=request_row, column=column_indices.get('usersAccepted', 0), 
                   value=0 if 'usersAccepted' in column_indices else None)
            ws.cell(row=request_row, column=column_indices.get('rejections', 0), 
                   value=json.dumps(rejections) if 'rejections' in column_indices else None)
            ws.cell(row=request_row, column=column_indices.get('lastStatusUpdate', 0), 
                   value=now.isoformat() if 'lastStatusUpdate' in column_indices else None)
            ws.cell(row=request_row, column=column_indices.get('lastStatusUpdateTime', 0), 
                   value=now.strftime("%H:%M:%S") if 'lastStatusUpdateTime' in column_indices else None)
            ws.cell(row=request_row, column=column_indices.get('statusChangedBy', 0), 
                   value=username if 'statusChangedBy' in column_indices else None)
        
        wb.save(REQUESTS_FILE)
        
        # Return updated request
        updated_request = rows_to_dicts(ws)[request_row - 2]
        
        # Process JSON fields
        for field in ['acceptedBy', 'departments', 'rejections', 'participantsCompleted']:
            if field in updated_request and updated_request[field] and isinstance(updated_request[field], str):
                try:
                    updated_request[field] = json.loads(updated_request[field])
                except:
                    updated_request[field] = []
        
        return updated_request
    except Exception as e:
        print(f"Error abandoning request: {str(e)}", file=sys.stderr)
        return None

def reject_request(request_id, username, reason=''):
    """Reject a request."""
    try:
        if not os.path.exists(REQUESTS_FILE):
            return None
        
        wb = openpyxl.load_workbook(REQUESTS_FILE)
        ws = wb.active
        columns = get_column_names(ws)
        column_indices = {name: idx+1 for idx, name in enumerate(columns)}
        
        # Find request row by ID
        request_row = None
        request_data = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if str(row[0].value) == str(request_id):
                request_row = row_idx
                request_data = {col: row[i].value for i, col in enumerate(columns)}
                break
        
        if not request_row or not request_data:
            return None
        
        now = datetime.now()
        
        # Add rejection record
        rejections_str = request_data.get('rejections', '[]')
        try:
            rejections = json.loads(rejections_str) if rejections_str else []
        except:
            rejections = []
        
        if not isinstance(rejections, list):
            rejections = []
        
        rejections.append({
            'username': username,
            'reason': reason,
            'date': now.strftime("%d/%m/%Y %H:%M:%S")
        })
        
        # Update cells
        ws.cell(row=request_row, column=column_indices.get('status', 0), 
               value='Rejected' if 'status' in column_indices else None)
        ws.cell(row=request_row, column=column_indices.get('rejections', 0), 
               value=json.dumps(rejections) if 'rejections' in column_indices else None)
        ws.cell(row=request_row, column=column_indices.get('lastStatusUpdate', 0), 
               value=now.isoformat() if 'lastStatusUpdate' in column_indices else None)
        ws.cell(row=request_row, column=column_indices.get('lastStatusUpdateTime', 0), 
               value=now.strftime("%H:%M:%S") if 'lastStatusUpdateTime' in column_indices else None)
        ws.cell(row=request_row, column=column_indices.get('statusChangedBy', 0), 
               value=username if 'statusChangedBy' in column_indices else None)
        
        wb.save(REQUESTS_FILE)
        
        # Return updated request
        updated_request = rows_to_dicts(ws)[request_row - 2]
        
        # Process JSON fields
        for field in ['acceptedBy', 'departments', 'rejections', 'participantsCompleted']:
            if field in updated_request and updated_request[field] and isinstance(updated_request[field], str):
                try:
                    updated_request[field] = json.loads(updated_request[field])
                except:
                    updated_request[field] = []
        
        return updated_request
    except Exception as e:
        print(f"Error rejecting request: {str(e)}", file=sys.stderr)
        return None

def get_user_requests(username):
    """Get requests for a specific user."""
    try:
        all_requests = get_requests()
        user_requests = []
        
        for request in all_requests:
            # Check if user is creator
            if request.get('creator') == username:
                user_requests.append(request)
                continue
            
            # Check if user is in acceptedBy
            accepted_by = request.get('acceptedBy', [])
            if isinstance(accepted_by, list) and username in accepted_by:
                user_requests.append(request)
                continue
        
        return user_requests
    except Exception as e:
        print(f"Error getting user requests: {str(e)}", file=sys.stderr)
        return []

def filter_requests(filters_json):
    """Filter requests based on criteria."""
    try:
        filters = json.loads(filters_json)
        all_requests = get_requests()
        filtered_requests = []
        
        for request in all_requests:
            include = True
            
            # Apply filters
            for field, value in filters.items():
                if field == 'department' and value:
                    if request.get('department') != value:
                        include = False
                        break
                
                elif field == 'status' and value and value != 'All':
                    if request.get('status') != value:
                        include = False
                        break
                
                elif field == 'type' and value:
                    if request.get('type') != value:
                        include = False
                        break
                
                elif field == 'multiDepartment' and value:
                    multi_department = request.get('multiDepartment')
                    if multi_department not in ['TRUE', 'True', 'true', True, 1]:
                        include = False
                        break
                
                elif field == 'search' and value:
                    search_value = value.lower()
                    title = str(request.get('title', '')).lower()
                    description = str(request.get('description', '')).lower()
                    department = str(request.get('department', '')).lower()
                    creator = str(request.get('creator', '')).lower()
                    
                    if (search_value not in title and search_value not in description and 
                        search_value not in department and search_value not in creator):
                        include = False
                        break
            
            if include:
                filtered_requests.append(request)
        
        return filtered_requests
    except Exception as e:
        print(f"Error filtering requests: {str(e)}", file=sys.stderr)
        return []

def check_expired_requests():
    """Check and update expired requests."""
    try:
        if not os.path.exists(REQUESTS_FILE):
            return {'updated': False}
        
        wb = openpyxl.load_workbook(REQUESTS_FILE)
        ws = wb.active
        columns = get_column_names(ws)
        column_indices = {name: idx+1 for idx, name in enumerate(columns)}
        
        now = datetime.now()
        updated = False
        expired_count = 0
        archived_count = 0
        
        # Process each row
        rows_to_delete = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            request_data = {col: row[i].value for i, col in enumerate(columns)}
            request_id = request_data.get('id')
            
            # Skip if no ID
            if not request_id:
                continue
            
            # Check completed or rejected requests
            if request_data.get('status') in ['Completed', 'Rejected'] and request_data.get('lastStatusUpdate'):
                try:
                    status_date = datetime.fromisoformat(request_data.get('lastStatusUpdate'))
                    one_day_later = datetime.fromtimestamp(status_date.timestamp() + (24 * 60 * 60))
                    
                    if now > one_day_later and not request_data.get('isExpired'):
                        # Mark as expired
                        ws.cell(row=row_idx, column=column_indices.get('isExpired', 0), 
                               value=True if 'isExpired' in column_indices else None)
                        updated = True
                    
                    if request_data.get('isExpired') in ['TRUE', 'True', 'true', True, 1]:
                        # Delete expired request
                        rows_to_delete.append(row_idx)
                        updated = True
                        expired_count += 1
                except:
                    pass
            
            # Check pending requests
            if request_data.get('status') == 'Pending':
                created_date = None
                try:
                    if request_data.get('createdAt'):
                        created_date = datetime.fromisoformat(request_data.get('createdAt'))
                    elif request_data.get('dateCreated'):
                        created_date = datetime.strptime(request_data.get('dateCreated'), "%d/%m/%Y")
                except:
                    continue
                
                if not created_date:
                    continue
                
                # Set expiry days based on request type
                expiry_days = 30  # Default for regular requests
                
                if request_data.get('type') == 'project':
                    expiry_days = 60  # Projects get 60 days
                elif request_data.get('multiDepartment') in ['TRUE', 'True', 'true', True, 1]:
                    expiry_days = 45  # Multi-department requests get 45 days
                
                expiry_date = datetime.fromtimestamp(created_date.timestamp() + (expiry_days * 24 * 60 * 60))
                
                if request_data.get('type') == 'project':
                    # Projects get archived after expiry
                    if now > expiry_date and request_data.get('archived') not in ['TRUE', 'True', 'true', True, 1]:
                        ws.cell(row=row_idx, column=column_indices.get('archived', 0), 
                               value=True if 'archived' in column_indices else None)
                        ws.cell(row=row_idx, column=column_indices.get('archivedAt', 0), 
                               value=now.isoformat() if 'archivedAt' in column_indices else None)
                        updated = True
                        archived_count += 1
                    
                    # Archived projects get deleted after 7 days
                    if (request_data.get('archived') in ['TRUE', 'True', 'true', True, 1] and 
                        request_data.get('archivedAt')):
                        try:
                            archived_date = datetime.fromisoformat(request_data.get('archivedAt'))
                            delete_date = datetime.fromtimestamp(archived_date.timestamp() + (7 * 24 * 60 * 60))
                            
                            if now > delete_date:
                                rows_to_delete.append(row_idx)
                                updated = True
                                expired_count += 1
                        except:
                            pass
                else:
                    # Regular requests get deleted after expiry
                    if now > expiry_date:
                        rows_to_delete.append(row_idx)
                        updated = True
                        expired_count += 1
        
        # Delete rows in reverse order to avoid index shifting
        for row_idx in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_idx)
        
        if updated:
            wb.save(REQUESTS_FILE)
        
        return {
            'updated': updated,
            'expired_count': expired_count,
            'archived_count': archived_count
        }
    except Exception as e:
        print(f"Error checking expired requests: {str(e)}", file=sys.stderr)
        return {'updated': False, 'error': str(e)}

def can_user_accept_request(request_id, username, department):
    """Check if user can accept a request."""
    try:
        all_requests = get_requests()
        target_request = None
        
        for request in all_requests:
            if str(request.get('id')) == str(request_id):
                target_request = request
                break
        
        if not target_request:
            return {'canAccept': False, 'reason': 'Request not found'}
        
        # Check if request is already accepted by user
        accepted_by = target_request.get('acceptedBy', [])
        if isinstance(accepted_by, list) and username in accepted_by:
            return {'canAccept': False, 'reason': 'Already accepted'}
        
        # Check request status
        if target_request.get('status') not in ['Pending', 'Rejected']:
            return {'canAccept': False, 'reason': f"Cannot accept request with status: {target_request.get('status')}"}
        
        # Check if multi-department request
        multi_department = target_request.get('multiDepartment')
        request_type = target_request.get('type', 'request')
        
        if (multi_department in ['TRUE', 'True', 'true', True, 1] or 
            request_type in ['project', 'Project']):
            # For multi-department or projects, check if the user's department is required
            departments = target_request.get('departments', [])
            if isinstance(departments, str):
                try:
                    departments = json.loads(departments)
                except:
                    departments = []
            
            if departments and isinstance(departments, list) and department not in departments:
                return {'canAccept': False, 'reason': 'Your department is not required for this request'}
        else:
            # For regular requests, check if department matches
            if department != target_request.get('department'):
                return {'canAccept': False, 'reason': 'Request is for a different department'}
        
        return {'canAccept': True}
    except Exception as e:
        print(f"Error checking if user can accept request: {str(e)}", file=sys.stderr)
        return {'canAccept': False, 'reason': 'Internal error'}

def main():
    if len(sys.argv) < 2:
        print('Usage: python excel_operations.py <operation> [args...]', file=sys.stderr)
        sys.exit(1)
    
    operation = sys.argv[1]
    
    try:
        if operation == 'get_departments':
            result = get_departments()
            print(json.dumps(result))
        
        elif operation == 'get_users':
            result = get_users()
            print(json.dumps(result))
        
        elif operation == 'login_user':
            if len(sys.argv) < 3:
                print('Missing username', file=sys.stderr)
                sys.exit(1)
            
            username = sys.argv[2]
            password = sys.argv[3] if len(sys.argv) > 3 else ''
            
            result = login_user(username, password)
            if result:
                print(json.dumps(result))
            else:
                print(json.dumps(None))
        
        elif operation == 'update_user':
            if len(sys.argv) < 4:
                print('Missing user ID or data', file=sys.stderr)
                sys.exit(1)
            
            user_id = sys.argv[2]
            user_data = sys.argv[3]
            
            result = update_user(user_id, user_data)
            if result:
                print(json.dumps(result))
            else:
                print(json.dumps(None))
        
        elif operation == 'get_requests':
            result = get_requests()
            print(json.dumps(result))
        
        elif operation == 'create_request':
            if len(sys.argv) < 3:
                print('Missing request data', file=sys.stderr)
                sys.exit(1)
            
            request_data = sys.argv[2]
            
            result = create_request(request_data)
            if result:
                print(json.dumps(result))
            else:
                print(json.dumps(None))
        
        elif operation == 'update_request':
            if len(sys.argv) < 4:
                print('Missing request ID or data', file=sys.stderr)
                sys.exit(1)
            
            request_id = sys.argv[2]
            request_data = sys.argv[3]
            
            result = update_request(request_id, request_data)
            if result:
                print(json.dumps(result))
            else:
                print(json.dumps(None))
        
        elif operation == 'delete_request':
            if len(sys.argv) < 3:
                print('Missing request ID', file=sys.stderr)
                sys.exit(1)
            
            request_id = sys.argv[2]
            
            result = delete_request(request_id)
            print(json.dumps({'success': result}))
        
        elif operation == 'accept_request':
            if len(sys.argv) < 4:
                print('Missing request ID or username', file=sys.stderr)
                sys.exit(1)
            
            request_id = sys.argv[2]
            username = sys.argv[3]
            
            result = accept_request(request_id, username)
            if result:
                print(json.dumps(result))
            else:
                print(json.dumps(None))
        
        elif operation == 'complete_request':
            if len(sys.argv) < 4:
                print('Missing request ID or username', file=sys.stderr)
                sys.exit(1)
            
            request_id = sys.argv[2]
            username = sys.argv[3]
            
            result = complete_request(request_id, username)
            if result:
                print(json.dumps(result))
            else:
                print(json.dumps(None))
        
        elif operation == 'abandon_request':
            if len(sys.argv) < 4:
                print('Missing request ID or username', file=sys.stderr)
                sys.exit(1)
            
            request_id = sys.argv[2]
            username = sys.argv[3]
            
            result = abandon_request(request_id, username)
            if result:
                print(json.dumps(result))
            else:
                print(json.dumps(None))
        
        elif operation == 'reject_request':
            if len(sys.argv) < 4:
                print('Missing request ID or username', file=sys.stderr)
                sys.exit(1)
            
            request_id = sys.argv[2]
            username = sys.argv[3]
            reason = sys.argv[4] if len(sys.argv) > 4 else ''
            
            result = reject_request(request_id, username, reason)
            if result:
                print(json.dumps(result))
            else:
                print(json.dumps(None))
        
        elif operation == 'get_user_requests':
            if len(sys.argv) < 3:
                print('Missing username', file=sys.stderr)
                sys.exit(1)
            
            username = sys.argv[2]
            
            result = get_user_requests(username)
            print(json.dumps(result))
        
        elif operation == 'filter_requests':
            if len(sys.argv) < 3:
                print('Missing filters', file=sys.stderr)
                sys.exit(1)
            
            filters = sys.argv[2]
            
            result = filter_requests(filters)
            print(json.dumps(result))
        
        elif operation == 'check_expired_requests':
            result = check_expired_requests()
            print(json.dumps(result))
        
        elif operation == 'can_user_accept_request':
            if len(sys.argv) < 5:
                print('Missing parameters', file=sys.stderr)
                sys.exit(1)
            
            request_id = sys.argv[2]
            username = sys.argv[3]
            department = sys.argv[4]
            
            result = can_user_accept_request(request_id, username, department)
            print(json.dumps(result))
        
        else:
            print(f'Unknown operation: {operation}', file=sys.stderr)
            sys.exit(1)
    
    except Exception as e:
        print(f'Error: {str(e)}', file=sys.stderr)
        sys.exit(1)

if __name__ == '__main__':
    main()
